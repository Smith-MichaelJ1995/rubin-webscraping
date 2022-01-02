import openpyxl
import xlsxwriter
import pandas as pd
from selenium import webdriver 
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.common.by import By 
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC 
from pathlib import Path


class Phase0:

    def __init__(self, keywords):  

        # location codes file path
        locationCodesFilePath = "location-codes-datatable.xlsx" 

        # stage placeholder for location codes
        codes = []

        # fetch location codes file from system, convert into iterable object
        xlsx_file = Path("../location-codes/{}".format(locationCodesFilePath))
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        # traverse through all rows in location codes spreadsheet
        for row in sheet.iter_rows(2, sheet.max_row):

            # extract key values from row
            cellLocation = row[1]
            cellCountry = row[2]
            cellCode = row[3]

            # only record of country code is US
            if "United States" in cellLocation.value and cellCountry.value == "US":

                # determine if location code is a state
                if cellLocation.value.count(",") == 1:
                    codes.append(cellCode.value)

        # placeholder for all search links
        links = []

        # section for generating links, traverse through each term
        for keyword in keywords:

            # traverse through all terms
            for code in codes:

                # stage combination of keyword + link
                links.append(
                    "https://www.linkedin.com/search/results/people/?geoUrn=%5B%22{}%22%5D&keywords={}&origin=FACETED_SEARCH".format(code, keyword)
                )

        # creation of worksheet
        wb = xlsxwriter.Workbook("../output/linkedin-query-links.xlsx")
        ws = wb.add_worksheet()
        row = 0

        # stage all links into worksheet
        for link in links:
            ws.write(row, 0, link)
            row += 1

        # close worksheet
        wb.close()

    # build locale spreadsheet from provided file
    # locale codes path
    def generate_locale_file_from_system(self, path):

        # placeholder for data to be added to the speadsheet
        finalResults = []

        # fetch locale codes from filesystem
        localCodesRawFile = open(path, "r")

        # iterate through all lines in the file
        for line in localCodesRawFile:

            # create columns for: "location", "country", "code"
            rawData = line.split("\t")

            # remove newline character from "code"
            rawData[2] = rawData[2].split("\n")[0]

            # stage record for insertion into spreadsheet
            finalResults.append(rawData)

        # write to spreadsheet
        df = pd.DataFrame(data=finalResults, columns=['Location', 'Country', 'Code'])

        # create excel writer object
        writer = pd.ExcelWriter("location-codes-datatable.xlsx")

        # write dataframe to excel
        df.to_excel(writer)

        # save the excel
        writer.save()